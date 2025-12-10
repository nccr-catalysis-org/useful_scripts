#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Nov 26 16:44:53 2025

@author: nr
"""
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Sep 2 14:35:40 2025

@author: nr
"""
import argparse
import logging
import os
import re
import shutil as sh
from typing import Any, Callable, Dict, List, Match, Optional, Pattern, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter

# --- Logger Setup ---
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO) 
handler = logging.StreamHandler()
formatter = logging.Formatter('%(levelname)s:%(message)s')
handler.setFormatter(formatter)
if logger.handlers:
    for h in logger.handlers:
        logger.removeHandler(h)
logger.addHandler(handler)
# --- End Logger Setup ---

# Regex to find cell references in formulas (e.g., A1, B2, or Sheet1!A1)
# 1. Sheet Name part (optional, group 1): (?:'([^']+)'!)? OR ([A-Za-z0-9_]+!)?
#    - We use the simpler version here: capture (SheetName!)? or (CellRef)
CELL_REF_REGEX: Pattern[str] = re.compile(r"((?:'[^']+'!)?|(?:\w+!)?)([A-Z]+)(\d+)")

class InvalidFileFormatError(ValueError):
    """Raised when the file content does not match the expected format or schema."""
    pass


def get_padding_info(worksheet) -> Tuple[int, int]:
    """
    Returns the number of rows and columns to unpad from a sheet.
    This function handles horizontal and vertical padding independently.
    
    Returns: (rows_to_delete, cols_to_delete)
    """
    rows_to_delete = 0
    cols_to_delete = 0

    # 1. Find the first non-empty row (Vertical Padding)
    for row_idx, row in enumerate(worksheet.iter_rows()):
        # row_idx is 0-indexed here
        if any(cell.value is not None for cell in row):
            rows_to_delete = row_idx
            break
        # Optimization: If we checked the first 20 rows and all are empty, stop early.
        # This prevents looping max_row times if the sheet is mostly empty.
        if row_idx > 20 and rows_to_delete == 0: 
             break

    # 2. Find the first non-empty column (Horizontal Padding)
    # Iterate through columns from 1 to max_column
    for col_idx in range(1, worksheet.max_column + 1):
        # Check if the column is entirely empty
        is_col_empty = True
        # Check only the first max_row (or a reasonable limit)
        row_limit = min(worksheet.max_row, rows_to_delete + 50)
        for row_idx in range(rows_to_delete + 1, row_limit + 1): # Start search below padded rows
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                is_col_empty = False
                break
        
        if not is_col_empty:
            cols_to_delete = col_idx - 1
            break
            
    return rows_to_delete, cols_to_delete


def update_cross_sheet_formula(
    formula: str, 
    source_sheet_name: str, 
    all_sheets_padding_map: Dict[str, Dict[str, Any]]
) -> str:
    """
    Rewrites a formula string based on padding removals across all sheets.
    This function handles both internal (same sheet) and external (cross-sheet) references.
    
    Args:
        formula: The original formula string (e.g., '=SUM(A1:B10)' or '=Sheet2!A1').
        source_sheet_name: The name of the sheet containing the formula.
        all_sheets_padding_map: Global map of padding (rows/cols) deleted for every sheet.

    Returns:
        The updated formula string.
    """
    if formula is None:
        return formula

    def replace_cell_ref(match: Match[str]) -> str:
        """Callback function for the regex substitute."""
        
        # Group 1: Sheet reference (e.g., 'Sheet2!' or empty string for same-sheet)
        # Group 2: Column reference (e.g., A)
        # Group 3: Row reference (e.g., 1)
        sheet_ref, col_ref, row_ref = match.groups()
        
        # 1. Determine which sheet's padding map to use
        target_sheet_name = source_sheet_name # Default to the sheet containing the formula
        
        # If sheet_ref exists, it's a cross-sheet reference
        if sheet_ref:
            # Extract the sheet name from the reference (remove quotes, exclamation mark)
            # Examples: 'Sheet 2'! => 'Sheet 2', Sheet3! => Sheet3
            target_sheet_name = sheet_ref.strip("'!").strip()
            
            # If the target sheet doesn't exist in the map (e.g., it was deleted 
            # or is an external link), we cannot apply a correction.
            if target_sheet_name not in all_sheets_padding_map:
                return match.group(0) # Return original reference
        
        # 2. Get the padding for the TARGET sheet
        padding_info = all_sheets_padding_map.get(target_sheet_name, {'rows': 0, 'cols': 0})
        rows_to_delete = padding_info['rows']
        cols_to_delete = padding_info['cols']
        
        # If the target sheet had no padding, return original reference
        if rows_to_delete == 0 and cols_to_delete == 0:
            return match.group(0)

        # 3. Apply the shift calculation
        original_row = int(row_ref)
        original_col_idx = column_index_from_string(col_ref)
        
        # New row index (subtract padding)
        new_row = original_row - rows_to_delete
        # New column index (subtract padding)
        new_col_idx = original_col_idx - cols_to_delete

        # If the new reference is <= 0 or outside A1, the formula is likely invalid.
        if new_row <= 0 or new_col_idx <= 0:
            logger.warning(
                f"Formula in {source_sheet_name} references cell {col_ref}{row_ref} "
                f"in target sheet {target_sheet_name}. Deletion shifts it outside A1 (to {new_col_idx}{new_row}). "
                "Returning original reference for manual check."
            )
            return match.group(0) # Return original cell reference (e.g., "Sheet2!A1")

        new_col_ref = get_column_letter(new_col_idx)
        
        # Reconstruct the reference using the original sheet reference string
        return f"{sheet_ref}{new_col_ref}{new_row}"

    # Use the regex to find all cell references and replace them using the callback
    return CELL_REF_REGEX.sub(replace_cell_ref, formula)


def process_xlsx_file(filename: str, outname: str, unpad: bool, strip_text: bool) -> bool:
    """
    Core function to process an Excel file:
    1. Determines padding for all sheets.
    2. Performs text stripping and physical padding deletion.
    3. Rewrites formulas based on padding. (FIXED: Moved to a dedicated step after deletions)
    
    Args:
        filename: Path to the source file.
        outname: Path to save the processed file.
        unpad: If True, padding rows/cols are deleted.
        strip_text: If True, all text cell values are stripped.
        
    Returns:
        True if successful, False otherwise.
    """
    if not os.path.exists(filename):
        logger.error(f"File not found: {filename}")
        return False
        
    logger.info(f"Processing: {os.path.basename(filename)} (Unpad: {unpad}, Strip: {strip_text})")
    
    # 1. Load the workbook
    try:
        wb = load_workbook(filename)
    except Exception as e:
        logger.error(f"Error loading workbook {filename}: {e}")
        sh.copy(filename, outname) # Copy source to destination for safety
        return False

    # This map stores the original padding data for *all* sheets
    all_sheets_padding_map: Dict[str, Dict[str, Any]] = {}
    
    # 2. First Pass: Determine Padding for ALL Sheets (if requested)
    if unpad:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_to_delete, cols_to_delete = get_padding_info(ws)
            all_sheets_padding_map[sheet_name] = {'rows': rows_to_delete, 'cols': cols_to_delete}
            if rows_to_delete > 0 or cols_to_delete > 0:
                logger.info(f"  Sheet '{sheet_name}': Found {rows_to_delete} padding row(s), {cols_to_delete} padding col(s).")
    else:
        # Dummy map if not unpadding
        for sheet_name in wb.sheetnames:
            all_sheets_padding_map[sheet_name] = {'rows': 0, 'cols': 0}


    # 3. Second Pass: Strip Text and Apply Physical Unpadding
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        padding_info = all_sheets_padding_map[sheet_name]
        
        # Calculate start indices *after* padding deletion (if unpad is False, these are 1)
        # Note: We still iterate over the whole sheet to find text to strip, 
        # but the strip operation only happens before row/col deletion.
        # However, for efficiency, we can focus the iteration on non-padded areas.
        
        # --- Text Stripping (before deletion) ---
        if strip_text:
            # Only iterate over the part of the sheet that is NOT padding, 
            # to avoid wasting time on empty cells in padding rows/cols.
            start_row_for_strip = padding_info['rows'] + 1
            start_col_for_strip = padding_info['cols'] + 1
            
            for row in ws.iter_rows(min_row=start_row_for_strip, min_col=start_col_for_strip):
                for cell in row:
                    # Apply stripping only to text values (data_type 's')
                    if cell.data_type == 's' and isinstance(cell.value, str):
                        original_value = cell.value
                        stripped_value = original_value.strip()
                        if stripped_value != original_value:
                            cell.value = stripped_value

        # --- Apply Unpadding (Row/Column Deletion) ---
        rows_to_delete = padding_info['rows']
        cols_to_delete = padding_info['cols']

        if rows_to_delete > 0:
            # openpyxl delete_rows(idx, amount). idx=1 means delete from the start.
            ws.delete_rows(1, rows_to_delete)
        if cols_to_delete > 0:
            ws.delete_cols(1, cols_to_delete)

    
    # 4. Third Pass (FIX): Rewrite Formulas (after deletion and using the global map)
    if unpad:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # After deletion, iteration starts from 1,1
            for row in ws.iter_rows():
                for cell in row:
                    # Only process formula cells ('f' for formula)
                    if cell.data_type == 'f' and isinstance(cell.value, str): 
                        # Crucially, we pass the GLOBAL padding map here!
                        cell.value = update_cross_sheet_formula(
                            cell.value, 
                            sheet_name, 
                            all_sheets_padding_map
                        )

    # 5. Save the modified workbook
    try:
        wb.save(outname)
        logger.info(f"Successfully processed {os.path.basename(filename)}.")
        return True
    except Exception as e:
        logger.error(f"Error saving file {os.path.basename(outname)}: {e}")
        return False

def get_padding_info_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Returns padding info of a DataFrame

    Args
        df : the dataframe to unpad

    Returns:
        the unpadded DataFrame
    

    """
    x, y = 0, 0
    while df.iloc[x].isna().all():
        x += 1
    while df.iloc[:, y].isna().all():
        y += 1
    return x, y

def unpad_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Unpads a DataFrame

    Args
        df : the dataframe to unpad

    Returns:
        the unpadded dataframe
    

    """
    x, y = get_padding_info_df(df)
    return df.iloc[x:, y:]

def strip_text_df(df: pd.DataFrame) -> pd.DataFrame:
    t = df.copy()
    stripvalue = lambda x: x if not isinstance(x, str) else x.strip()
    return t.map(stripvalue)

def process_xls_file(filename: str, outname: str, unpad: bool, strip_text: bool) -> None:
    if not os.path.exists(filename):
        logger.error(f"File not found: {filename}")
        return False
        
    logger.info(f"Processing: {os.path.basename(filename)} (Unpad: {unpad}, Strip: {strip_text})")
    
    try:
        dfs = pd.read_excel(filename, sheet_name=None)
    except Exception as e:
        logger.error(f"Error loading file {filename}: {e}")
        sh.copy(filename, outname) # Copy source to destination for safety
        return False
    for name, df in dfs.items():
        if unpad:
            df = unpad_df(df)
        if strip_text:
            df = strip_text_df(df)
        dfs[name] = df
    with pd.ExcelWriter(outname, engine='xlwt') as writer:
        for name, df in dfs.items():
            df.to_excel(writer, sheet_name=name, index=False)

def process_csv_file(filename: str, outname: str, unpad: bool, strip_text: bool) -> None:
    if not os.path.exists(filename):
        logger.error(f"File not found: {filename}")
        return False
        
    logger.info(f"Processing: {os.path.basename(filename)} (Unpad: {unpad}, Strip: {strip_text})")
    
    try:
        df = pd.read_csv(filename)
    except Exception as e:
        logger.error(f"Error loading file {filename}: {e}")
        sh.copy(filename, outname) # Copy source to destination for safety
        return False
    if unpad:
        df = unpad_df(df)
    if strip_text:
        df = strip_text_df(df)
    df.to_csv(outname, index=False)

def process_folder(source_fol: str, dest_fol: str, unpad: bool, strip_text: bool):
    """Recursively processes all Excel files in a folder."""
    source_fol = os.path.abspath(source_fol)
    dest_fol = os.path.abspath(dest_fol)
    
    if not os.path.exists(source_fol):
        logger.error(f"Source folder not found: {source_fol}")
        return

    os.makedirs(dest_fol, exist_ok=True)
    
    logger.info(f"--- Starting Folder Process: {os.path.basename(source_fol)} to {os.path.basename(dest_fol)}/ ---")

    for fol, subfols, files in os.walk(source_fol):
        # Create corresponding destination folder
        destfol = fol.replace(source_fol, dest_fol)
        for subfol in subfols:
            os.makedirs(os.path.join(destfol, subfol), exist_ok=True)
        
        for file in files:
            source_path = os.path.join(fol, file)
            dest_path = os.path.join(destfol, file)
            
            if file.lower().endswith(".xlsx"):
                # Process Excel files
                process_xlsx_file(source_path, dest_path, unpad, strip_text)
            else:
                if file.lower().endswith(".xls"):
                    process_xls_file(source_path, dest_path, unpad, strip_text)
                elif file.lower().endswith(".csv"):
                    process_csv_file(source_path, dest_path, unpad, strip_text)
                else:
                    # Copy other files directly
                    try:
                        sh.copy2(source_path, dest_path)
                    except Exception as e:
                        logger.error(f"Error copying non-Excel file {file}: {e}")
            

    logger.info(f"--- Folder Process Complete: {os.path.basename(dest_fol)} ---")

def read_sheets(file, frmt=None):
    """
    Reads an Excel (xlsx/xls) or CSV file into a pandas DataFrame.
    """
    file_lower = file.lower()
    if frmt is None:
        if file_lower.endswith(".csv"):
            frmt = "csv"
        elif file_lower.endswith(".xlsx"):
            frmt = "xlsx"
        elif file_lower.endswith(".xls"): # Added XLS support
            frmt = "xls"
        else:
            raise InvalidFileFormatError("Only csv, xlsx, and xls can be processed")
    
    # Determine the read function
    if frmt == "csv":
        # Read csv
        sheets = {"csv_only_sheet": {"df": pd.read_csv(file)}}
        # Re-read the file to capture the actual first row as the list of column names
        try:
            # Set header=None and read only the first row
            columns_df = pd.read_csv(file, nrows=1, header=None)
            # Extract the original column names (potentially non-unique)
            sheets["csv_only_sheet"]["columns"] = columns_df.iloc[0].fillna(value="").astype(str).tolist()
        except Exception as e:
            logger.warning(f"Could not read first row for original column names in {file}: {e}")
            sheets["csv_only_sheet"]["columns"] = sheets["csv_only_sheet"]["df"].columns.astype(str).tolist() # Fallback to pandas detected headers
    elif frmt in ["xlsx", "xls"]:
        sheets = {k: {"df": v} for k, v in pd.read_excel(file, sheet_name=None).items()}
        for sheet_name, dict_ in sheets.items():
            try:
                # Set header=None and read only the first row
                columns_df = pd.read_excel(file, nrows=1, header=None, sheet_name=sheet_name)
                # Extract the original column names (potentially non-unique)
                sheets[sheet_name]["columns"] = columns_df.iloc[0].fillna(value="").astype(str).tolist()
            except Exception as e:
                logger.warning(f"Could not read first row for original column names in {file}: {e}")
                sheets[sheet_name]["columns"] = sheets[sheet_name]["table"].columns.astype(str).tolist() # Fallback to pandas detected headers
    else:
         raise InvalidFileFormatError(f"Unsupported format for reading: {frmt}")

    return sheets, frmt

def check_for_multiple_tables(df, file):
    """
    Checks if a DataFrame contains multiple tables separated by fully empty columns (NaNs).
    Returns True if multiple tables are found, False otherwise.
    """
    empty_cols = []
    
    for n, col in enumerate(df.columns):
        # Check if the entire column consists of NaN values
        if df[col].isna().all():
            empty_cols.append(n)

    if empty_cols:
        logger.warning(f"Multiple tables issue in {file}: Columns at indices {empty_cols} are empty and suggest a table split.")
        return True
    return False

def _safe_sheet_name(name: Any) -> str:
    """Sanitizes a string to be a valid, truncated Excel sheet name."""
    name_str = str(name)
    # Invalid characters: \ / ? * [ ] : 
    name_str = name_str.replace('[', '').replace(']', '').replace(':', '').replace('/', '').replace('\\', '').replace('?', '').replace('*', '').replace(' ', '_')
    # Limit length to 31 characters
    return name_str[:31]

def _get_excel_writer_engine(out_format: str) -> str:
    """Returns the correct Pandas ExcelWriter engine based on output format."""
    if out_format == 'xlsx':
        return 'xlsxwriter'
    elif out_format == 'xls':
        # Requires the 'xlwt' library
        return 'xlwt'
    else:
        raise InvalidFileFormatError(f"Unsupported Excel format for writing: {out_format}")

def split_tables(file, in_format=None, out_format=None, inplace=False, destination=None):
    """
    Splits a file containing multiple tables separated by empty columns into 
    individual sheets (if XLSX/XLS) or separate CSV files.
    """
    try:
        sheets, in_format = read_sheets(file, frmt=in_format)
    except InvalidFileFormatError as e:
        logger.error(f"Skipping split for {file}: {e}")
        return
    
    tables_per_sheet = {}
    for sheet_name, data in sheets.items():
        df, original_columns = data["df"], data["columns"]
        start = 0
        tables: Dict[str, pd.DataFrame] = {}
        
        # 1. Identify table boundaries (columns that are all NaN)
        empty_cols_indices = [n for n, col in enumerate(df.columns) if df[col].isna().all()]
        table_boundaries = empty_cols_indices + [len(df.columns)]
        
        # 2. Extract tables
        for n, block_end in enumerate(table_boundaries):
            end = block_end
            if end > start:  # Extract table if block is not empty
                table = df.iloc[:, start:end].copy()
                if not any(original_columns[start:end]):
                    table = table.iloc[1:].reset_index(drop=True)
                    key = start
                elif original_columns[start] and not any(original_columns[start+1:end]):  # table title only
                    key = df.columns[start]
                    table.columns = table.iloc[0].tolist()
                    table = table.iloc[1:].reset_index(drop=True)  # perhaps redundant as we do not write index, but still cleaner in case of future development
                else:
                    table.columns = original_columns[start:end]
                    key = start
                tables[key] = table
            start = block_end + 1
        if not tables:
            logger.info(f"No tables found in sheet {sheet_name} in {file} to split.")
            tables_per_sheet[sheet_name] = {sheet_name: df}
        tables_per_sheet[sheet_name] = tables

    # 3. Handle Output Path
    out_format = in_format if out_format is None else out_format
    basename = os.path.splitext(file)[0]
    
    if out_format in ['xlsx', 'xls']:
        if destination:
            outfile = destination
        elif inplace:
            outfile = file
        else:
            outfile = f"{basename}_split.{out_format}"
        
        # Ensure the output directory exists
        os.makedirs(os.path.dirname(outfile) or '.', exist_ok=True)
        
        # Ensure the filename is not too long
        if len(outfile) > 218: 
             logger.warning(f"Output filename {outfile} is too long, truncating.")
             outfile = f"{basename[:210]}_split.{out_format}"

        # 4. Write Output (XLSX/XLS)
        try:
            engine = _get_excel_writer_engine(out_format)
            with pd.ExcelWriter(outfile, engine=engine) as writer:
                for sheet_name, tables in tables_per_sheet.items():
                    for k, table in tables.items():
                        new_sheet_name = _safe_sheet_name(f"{sheet_name}_{k}")
                        table.to_excel(writer, sheet_name=new_sheet_name, index=False, header=True)
            logger.info(f"Successfully split tables from {file} to {out_format} sheets in {outfile}")
        except Exception as e:
            logger.error(f"Error writing {out_format} output for {file}: {e}")

    elif out_format == "csv":
        # Write each table to a separate CSV file
        for k, table in tables_per_sheet["csv_only_sheet"].items():
            safe_k = _safe_sheet_name(k)
            outfile = os.path.join(f"{basename}_split_{safe_k}.csv")
            table.to_csv(outfile, index=False, header=True)
        logger.info(f"Successfully split tables from {file} into multiple CSV")
    else:
        logger.error(f"Unsupported output format: {out_format}")


def split_into_two_colum_tables(file, in_format=None, out_format=None, inplace=False, destination=None):
    """
    Splits a file containing multiple tables into two-column (X, Y) pairs.
    """
    try:
        df, original_columns, in_format = read_table(file, frmt=in_format)
    except InvalidFileFormatError as e:
        logger.error(f"Skipping 2-column split for {file}: {e}")
        return

    tables: Dict[str, pd.DataFrame] = {}
    
    # 1. Identify table boundaries (columns that are all NaN)
    empty_cols_indices = [n for n, col in enumerate(df.columns) if df[col].isna().all()]
    
    # Define block start and end indices
    block_starts = [0] + [i + 1 for i in empty_cols_indices]
    block_ends = empty_cols_indices + [len(df.columns)]
    
    # Filter for valid blocks (start < end)
    blocks: List[Tuple[int, int]] = [(s, e) for s, e in zip(block_starts, block_ends) if s < e]

    # 2. Extract tables in (X, Y) pairs
    for start, end in blocks:
        block_df = df.iloc[:, start:end]
        
        # The first column of the block is the X-axis (start).
        if block_df.shape[1] > 1:
            x_col_index = start
            x_col_name_original = original_columns[x_col_index]
            
            # Create (X, Y) pairs for all Y columns in this block
            for y_index_in_block in range(1, block_df.shape[1]):
                y_col_index = start + y_index_in_block
                
                y_col_name_original = original_columns[y_col_index] 
                y_col_name_unique = df.columns[y_col_index] # Unique Pandas name for key
                
                # Create the two-column table
                table = df.iloc[:, [x_col_index, y_col_index]].copy()
                
                # Use original column names for the headers
                table.columns = [x_col_name_original, y_col_name_original]
                
                # Use the unique Pandas Y column name as the key
                key = y_col_name_unique
                tables[key] = table

    if not tables:
        logger.info(f"No two-column tables found in {file} to split.")
        return

    # 3. Handle Output Path
    out_format = in_format if out_format is None else out_format
    basename = os.path.splitext(file)[0]
    
    if out_format in ['xlsx', 'xls']:
        if destination:
            outfile = destination
        elif inplace:
            outfile = file
        else:
            outfile = f"{basename}_2col_split.{out_format}"
            
        os.makedirs(os.path.dirname(outfile) or '.', exist_ok=True)
        
        if len(outfile) > 218: 
             logger.warning(f"Output filename {outfile} is too long, truncating.")
             outfile = f"{basename[:210]}_2col_split.{out_format}"

        # 4. Write Output (XLSX/XLS)
        try:
            engine = _get_excel_writer_engine(out_format)
            with pd.ExcelWriter(outfile, engine=engine) as writer:
                for k, table in tables.items():
                    # k is the unique Pandas name (e.g., 'Y.1')
                    sheet_name = _safe_sheet_name(k)
                    table.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
            logger.info(f"Successfully split into two-column tables from {file} to {out_format} sheets in {outfile}")
        except Exception as e:
            logger.error(f"Error writing {out_format} output for {file}: {e}")

    elif out_format == "csv":
        # Write each table to a separate CSV file in a new directory
        output_dir = os.path.splitext(file)[0] + "_2col_split_csv" if destination is None else destination
        os.makedirs(output_dir, exist_ok=True)
        
        for k, table in tables.items():
            safe_k = _safe_sheet_name(k).lower()
            outfile = os.path.join(output_dir, f"{safe_k}.csv")
            table.to_csv(outfile, index=False, header=True)
        logger.info(f"Successfully split tables from {file} into multiple 2-column CSV files in directory: {output_dir}")
    else:
        logger.error(f"Unsupported output format: {out_format}")



# =============================================================================
# Checking function 
# =============================================================================

def check_xls_file(filename: str, check_padding: bool, check_strip: bool) -> Optional[Dict[str, Any]]:
    """
    Checks a single .xls file for padding and/or unstripped text.
    
    Returns:
        A dictionary of issues found, or None if no issues found or file is not Excel/error occurred.
    """
    if not filename.lower().endswith(".xlsx"):
        return None
        
    try:
        dfs = pd.read_excel(filename, sheet_name=None)
    except Exception:
        logger.error(f"Could not load file {filename}. Skipping check.")
        return None
        
    issues = {'padding_found': False, 'strip_issues': False, 'details': {}}
    
    for name, df in dfs.items():
        sheet_issues = {'padding': (0, 0), 'strip_cells': []}
        
        # 1. Check Padding
        if check_padding:
            rows_to_delete, cols_to_delete = get_padding_info_df(df)
            if rows_to_delete > 0 or cols_to_delete > 0:
                issues['padding_found'] = True
                sheet_issues['padding'] = (rows_to_delete, cols_to_delete)
        
        # 2. Check Text Stripping
        if check_strip:
            # We only need to check the remaining cells (i.e., skipping any padded area)
            start_row = sheet_issues['padding'][0] + 1
            start_col = sheet_issues['padding'][1] + 1
            
            for n, row in df.iloc[start_row:, start_col:].iter_rows():
                for m, cell in enumerate(row):
                    # Check for text (data_type 's') that starts or ends with space
                    if isinstance(cell, str):
                        if cell.value.startswith(' ') or cell.value.endswith(' '):
                            issues['strip_issues'] = True
                            sheet_issues['strip_cells'].append([n + start_row, m + start_col])
            
        if issues['padding_found'] or issues['strip_issues']:
            issues['details'][name] = sheet_issues
            
    return issues if issues['padding_found'] or issues['strip_issues'] else None

def check_xlsx_file(filename: str, check_padding: bool, check_strip: bool) -> Optional[Dict[str, Any]]:
    """
    Checks a single .xlsx file for padding and/or unstripped text.
    
    Returns:
        A dictionary of issues found, or None if no issues found or file is not Excel/error occurred.
    """
    if not filename.lower().endswith(".xlsx"):
        return None
        
    try:
        wb = load_workbook(filename)
    except Exception:
        logger.error(f"Could not load file {filename}. Skipping check.")
        return None
        
    issues = {'padding_found': False, 'strip_issues': False, 'details': {}}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_issues = {'padding': (0, 0), 'strip_cells': []}
        
        # 1. Check Padding 
        if check_padding:
            rows_to_delete, cols_to_delete = get_padding_info(ws)
            if rows_to_delete > 0 or cols_to_delete > 0:
                issues['padding_found'] = True
                sheet_issues['padding'] = (rows_to_delete, cols_to_delete)
        
        # 2. Check Text Stripping 
        if check_strip:
            # We only need to check the remaining cells (i.e., skipping any padded area)
            start_row = sheet_issues['padding'][0] + 1
            start_col = sheet_issues['padding'][1] + 1
            
            for row in ws.iter_rows(min_row=start_row, min_col=start_col):
                for cell in row:
                    # Check for text (data_type 's') that starts or ends with space
                    if cell.data_type == 's' and isinstance(cell.value, str):
                        if cell.value.startswith(' ') or cell.value.endswith(' '):
                            issues['strip_issues'] = True
                            sheet_issues['strip_cells'].append(cell.coordinate)
            
        if issues['padding_found'] or issues['strip_issues']:
            issues['details'][sheet_name] = sheet_issues
            
    return issues if issues['padding_found'] or issues['strip_issues'] else None

def check_csv_file(filename: str, check_padding: bool, check_strip: bool) -> Optional[Dict[str, Any]]:
    """
    Checks a single csv file for padding and/or unstripped text.
    
    Returns:
        A dictionary of issues found, or None if no issues found or file is not Excel/error occurred.
    """
    if not filename.lower().endswith(".csv"):
        return None
        
    try:
        df = pd.read_csv(filename)
    except Exception:
        logger.error(f"Could not load file {filename}. Skipping check.")
        return None
        
    issues = {'padding_found': False, 'strip_issues': False, 'details': {"only_sheet": {'padding': (0, 0), 'strip_cells': []}}}
    
    # 1. Check Padding
    if check_padding:
        rows_to_delete, cols_to_delete = get_padding_info_df(df)
        if rows_to_delete > 0 or cols_to_delete > 0:
            issues['padding_found'] = True
            issues["details"]["only_sheet"]['padding'] = (rows_to_delete, cols_to_delete)
    
    # 2. Check Text Stripping
    if check_strip:
        # We only need to check the remaining cells (i.e., skipping any padded area)
        start_row = issues["details"]["only_sheet"]['padding'][0] + 1
        start_col = issues["details"]["only_sheet"]['padding'][1] + 1
        
        for n, row in df.iloc[start_row:, start_col:].iterrows():
            for m, cell in enumerate(row):
                # Check for text (data_type 's') that starts or ends with space
                if isinstance(cell, str):
                    if cell.value.startswith(' ') or cell.value.endswith(' '):
                        issues['strip_issues'] = True
                        issues["details"]["only_sheet"]['strip_cells'].append([n + start_row, m + start_col])
        
    return issues if issues['padding_found'] or issues['strip_issues'] else None


def check_folder_recursively(folder_path: str, check_padding: bool, check_strip: bool):
    """
    Recursively checks all Excel files in a folder for issues and prints a report.
    """
    if not os.path.isdir(folder_path):
        logger.error(f"Folder not found: {folder_path}")
        return

    logger.info(f"--- Starting Recursive Check in: {os.path.basename(folder_path)} ---")
    
    found_issues = False
    
    for fol, _, files in os.walk(folder_path):
        for file in files:
            logger.debug(file)
            full_path = os.path.join(fol, file)
            
            if file.lower().endswith(".csv"):
                issues = check_csv_file(full_path, check_padding, check_strip)
            elif file.lower().endswith(".xlsx"):
                issues = check_xlsx_file(full_path, check_padding, check_strip)
            elif file.lower().endswith(".xls"):
                issues = check_xls_file(full_path, check_padding, check_strip)
            else:
                issues = False
            if issues:
                found_issues = True
                relative_path = os.path.relpath(full_path, folder_path)
                logger.warning(f"\n[ISSUE FOUND]: {relative_path}")
                
                if issues['padding_found']:
                    logger.warning("  * Padding Found (Run 'unpad' or 'clean' to fix):")
                    for sheet, detail in issues['details'].items():
                        if detail['padding'][0] > 0 or detail['padding'][1] > 0:
                            logger.warning(f"    - Sheet '{sheet}': {detail['padding'][0]} row(s), {detail['padding'][1]} col(s)")
                            
                if issues['strip_issues']:
                    logger.warning("  * Unstripped Text Found (Run 'strip-text' or 'clean' to fix):")
                    for sheet, detail in issues['details'].items():
                        if detail['strip_cells']:
                            # Only show the first few cells to keep the output clean
                            coords = detail['strip_cells'][:5]
                            more = f"... (+{len(detail['strip_cells']) - 5} more)" if len(detail['strip_cells']) > 5 else ""
                            logger.warning(f"    - Sheet '{sheet}': e.g., {', '.join(coords)}{more}")
                logger.debug("done")

    if not found_issues:
        logger.info("\n--- Check Complete: No issues found in any Excel file. ---")
    else:
        logger.info("\n--- Check Complete: Issues reported above. ---")


# =============================================================================
# CLI Implementation
# =============================================================================

def handle_process_command(args, unpad: bool, strip_text: bool):
    """
    Handler function for 'unpad', 'strip-text', and 'clean' commands.
    """
    logger.info(args)
    if not os.path.exists(args.source):
        logger.error(f"Source path not found: {args.source}")
        return

    if os.path.isdir(args.source):
        # Handle folder processing
        dest = args.destination if args.destination else args.source + "_processed" 
        process_folder(args.source, dest, unpad, strip_text)
    
    elif os.path.isfile(args.source):
        # Handle single file processing
        if not args.source.lower().endswith(".xlsx"):
            logger.error(f"File must be an .xlsx file: {args.source}")
            return
            
        dest = args.destination if args.destination else args.source.replace(".xlsx", "_processed.xlsx")
        process_xlsx_file(args.source, dest, unpad, strip_text)
        
    else:
        logger.error("Source must be a valid file or directory.")


def handle_check_command(args):
    """
    Handler function for the 'check' command.
    """
    check_padding = args.padding
    check_strip = args.strip
    
    # If neither is specified, check both by default
    if not check_padding and not check_strip:
        check_padding = True
        check_strip = True
        
    check_folder_recursively(args.folder, check_padding, check_strip)
    

def cli():
    """Configures and runs the command line interface."""
    parser = argparse.ArgumentParser(
        description=(
            "A utility for cleaning Excel files (unpadding and text stripping) "
            "and checking folders for common issues."
        ),
        formatter_class=argparse.RawTextHelpFormatter
    )
    subparsers = parser.add_subparsers(dest='command', required=True)

    # --- Parent Parser for Processing Commands (unpad, strip-text, clean) ---
    process_parser = argparse.ArgumentParser(add_help=False)
    process_parser.add_argument(
        'source',
        type=str,
        help='Path to the source .xlsx file OR directory to process.'
    )
    process_parser.add_argument(
        'destination',
        type=str,
        default=None,
        help='Optional path for the output file or directory. If a file is given, defaults to <file>_processed.xlsx. If a folder is given, defaults to <folder>_processed.'
    )

    # --- 'unpad' command parser ---
    parser_unpad = subparsers.add_parser(
        'unpad',
        help='Removes leading empty rows and columns from Excel sheets and updates formulas.',
        parents=[process_parser]
    )
    parser_unpad.set_defaults(func=lambda args: handle_process_command(args, unpad=True, strip_text=False))

    # --- 'strip-text' command parser ---
    parser_strip = subparsers.add_parser(
        'strip-text',
        help='Strips leading and trailing whitespace from all text cells.',
        parents=[process_parser]
    )
    parser_strip.set_defaults(func=lambda args: handle_process_command(args, unpad=False, strip_text=True))
    
    # --- 'clean' command parser (Unpad + Strip) ---
    parser_clean = subparsers.add_parser(
        'clean',
        help='Performs both unpadding and text stripping (recommended).',
        parents=[process_parser]
    )
    parser_clean.set_defaults(func=lambda args: handle_process_command(args, unpad=True, strip_text=True))
    
    # --- 'check' command parser (Q2) ---
    parser_check = subparsers.add_parser(
        'check',
        help='Recursively checks a folder for padding or unstripped text issues.',
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser_check.add_argument(
        'folder',
        type=str,
        help='Path to the root directory to check recursively.'
    )
    check_group = parser_check.add_mutually_exclusive_group()
    check_group.add_argument(
        '--padding',
        action='store_true',
        help='Only check for padding issues.'
    )
    check_group.add_argument(
        '--strip',
        action='store_true',
        help='Only check for unstripped text issues.'
    )
    parser_check.set_defaults(func=handle_check_command)

    # Parse the arguments and call the handler function
    args = parser.parse_args()
    args.func(args)


if __name__ == '__main__':
    cli()