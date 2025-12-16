#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Nov 26 16:44:53 2025

@author: nr
"""

import argparse
from itertools import product
import logging
import os
import re
import shutil as sh
import sys
from typing import Any, Callable, Dict, List, Match, Optional, Pattern, Tuple


import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter

# --- Logger Setup ---
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO) 
handler = logging.StreamHandler()
formatter = logging.Formatter('%(levelname)s:%(message)s')
handler.setFormatter(formatter)
if logger.handlers:
    for h in logger.handlers:
        logger.removeHandler(h)
logger.addHandler(handler)
# --- End Logger Setup ---

# Regex to find cell references in formulas (e.g., A1, B2, or Sheet1!A1)
# 1. Sheet Name part (optional, group 1): (?:'([^']+)'!)? OR ([A-Za-z0-9_]+!)?
#    - We use the simpler version here: capture (SheetName!)? or (CellRef)
CELL_REF_REGEX: Pattern[str] = re.compile(r"((?:'[^']+'!)?|(?:\w+!)?)([A-Z]+)(\d+)")
PROCESS_EXTENSIONS: Tuple[str, ...] = ("xlsx", "xls") 
STRICT_SEP_EXTENSIONS: Tuple[str, ...] = ("csv", "tsv")
WIDE_SEP_EXTENSIONS: Tuple[str, ...] = ("csv", "tsv", "txt", "dat")
EXT_TO_SEP = {"csv": ",", "tsv": "\t", "txt": "\s+", "dat": "\s+"}
TABULAR_EXTENSIONS: Tuple[str, ...] = PROCESS_EXTENSIONS + STRICT_SEP_EXTENSIONS

class InvalidFileFormatError(ValueError):
    """Raised when the file content does not match the expected format or schema."""
    pass

###############################################################################
# Openpyxel worksheet handling for unpadding and text stripping
###############################################################################

def get_padding_info_ws(worksheet) -> Tuple[int, int]:
    """
    Returns the number of rows and columns to unpad from a sheet.
    This function handles horizontal and vertical padding independently.
    
    Returns: (rows_to_delete, cols_to_delete)
    """
    rows_to_delete = 0
    cols_to_delete = 0

    # 1. Find the first non-empty row (Vertical Padding)
    for row_idx, row in enumerate(worksheet.iter_rows()):
        # row_idx is 0-indexed here
        if any(cell.value is not None for cell in row):
            rows_to_delete = row_idx
            break
        # Optimization: If we checked the first 20 rows and all are empty, stop early.
        # This prevents looping max_row times if the sheet is mostly empty.
        if row_idx > 20 and rows_to_delete == 0: 
             break

    # 2. Find the first non-empty column (Horizontal Padding)
    # Iterate through columns from 1 to max_column
    for col_idx in range(1, worksheet.max_column + 1):
        # Check if the column is entirely empty
        is_col_empty = True
        # Check only the first max_row (or a reasonable limit)
        row_limit = min(worksheet.max_row, rows_to_delete + 50)
        for row_idx in range(rows_to_delete + 1, row_limit + 1): # Start search below padded rows
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                is_col_empty = False
                break
        
        if not is_col_empty:
            cols_to_delete = col_idx - 1
            break
            
    return rows_to_delete, cols_to_delete


def update_cross_sheet_formula(
    formula: str, 
    source_sheet_name: str, 
    all_sheets_padding_map: Dict[str, Dict[str, Any]]
) -> str:
    """
    Rewrites a formula string based on padding removals across all sheets.
    This function handles both internal (same sheet) and external (cross-sheet) references.
    
    Args:
        formula: The original formula string (e.g., '=SUM(A1:B10)' or '=Sheet2!A1').
        source_sheet_name: The name of the sheet containing the formula.
        all_sheets_padding_map: Global map of padding (rows/cols) deleted for every sheet.

    Returns:
        The updated formula string.
    """
    if formula is None:
        return formula

    def replace_cell_ref(match: Match[str]) -> str:
        """Callback function for the regex substitute."""
        
        # Group 1: Sheet reference (e.g., 'Sheet2!' or empty string for same-sheet)
        # Group 2: Column reference (e.g., A)
        # Group 3: Row reference (e.g., 1)
        sheet_ref, col_ref, row_ref = match.groups()
        
        # 1. Determine which sheet's padding map to use
        target_sheet_name = source_sheet_name # Default to the sheet containing the formula
        
        # If sheet_ref exists, it's a cross-sheet reference
        if sheet_ref:
            # Extract the sheet name from the reference (remove quotes, exclamation mark)
            # Examples: 'Sheet 2'! => 'Sheet 2', Sheet3! => Sheet3
            target_sheet_name = sheet_ref.strip("'!").strip()
            
            # If the target sheet doesn't exist in the map (e.g., it was deleted 
            # or is an external link), we cannot apply a correction.
            if target_sheet_name not in all_sheets_padding_map:
                return match.group(0) # Return original reference
        
        # 2. Get the padding for the TARGET sheet
        padding_info = all_sheets_padding_map.get(target_sheet_name, {'rows': 0, 'cols': 0})
        rows_to_delete = padding_info['rows']
        cols_to_delete = padding_info['cols']
        
        # If the target sheet had no padding, return original reference
        if rows_to_delete == 0 and cols_to_delete == 0:
            return match.group(0)

        # 3. Apply the shift calculation
        original_row = int(row_ref)
        original_col_idx = column_index_from_string(col_ref)
        
        # New row index (subtract padding)
        new_row = original_row - rows_to_delete
        # New column index (subtract padding)
        new_col_idx = original_col_idx - cols_to_delete

        # If the new reference is <= 0 or outside A1, the formula is likely invalid.
        if new_row <= 0 or new_col_idx <= 0:
            logger.warning(
                f"Formula in {source_sheet_name} references cell {col_ref}{row_ref} "
                f"in target sheet {target_sheet_name}. Deletion shifts it outside A1 (to {new_col_idx}{new_row}). "
                "Returning original reference for manual check."
            )
            return match.group(0) # Return original cell reference (e.g., "Sheet2!A1")

        new_col_ref = get_column_letter(new_col_idx)
        
        # Reconstruct the reference using the original sheet reference string
        return f"{sheet_ref}{new_col_ref}{new_row}"

    # Use the regex to find all cell references and replace them using the callback
    return CELL_REF_REGEX.sub(replace_cell_ref, formula)


def unpad_strip_xlsx_file(filename: str, outname: str, unpad: bool, strip_text: bool) -> bool:
    """
    Core function to process an Excel file:
    1. Determines padding for all sheets.
    2. Performs text stripping and physical padding deletion.
    3. Rewrites formulas based on padding. (FIXED: Moved to a dedicated step after deletions)
    
    Args:
        filename: Path to the source file.
        outname: Path to save the processed file.
        unpad: If True, padding rows/cols are deleted.
        strip_text: If True, all text cell values are stripped.
        
    Returns:
        True if successful, False otherwise.
    """
    if not os.path.exists(filename):
        logger.error(f"File not found: {filename}")
        return False
        
    logger.info(f"Processing: {os.path.basename(filename)} (Unpad: {unpad}, Strip: {strip_text})")
    
    # 1. Load the workbook
    try:
        wb = load_workbook(filename)
    except Exception as e:
        logger.error(f"Error loading workbook {filename}: {e}")
        sh.copy(filename, outname) # Copy source to destination for safety
        return False

    # This map stores the original padding data for *all* sheets
    all_sheets_padding_map: Dict[str, Dict[str, Any]] = {}
    
    # 2. First Pass: Determine Padding for ALL Sheets (if requested)
    if unpad:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_to_delete, cols_to_delete = get_padding_info_ws(ws)
            all_sheets_padding_map[sheet_name] = {'rows': rows_to_delete, 'cols': cols_to_delete}
            if rows_to_delete > 0 or cols_to_delete > 0:
                logger.info(f"  Sheet '{sheet_name}': Found {rows_to_delete} padding row(s), {cols_to_delete} padding col(s).")
    else:
        # Dummy map if not unpadding
        for sheet_name in wb.sheetnames:
            all_sheets_padding_map[sheet_name] = {'rows': 0, 'cols': 0}


    # 3. Second Pass: Strip Text and Apply Physical Unpadding
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        padding_info = all_sheets_padding_map[sheet_name]
        
        # Calculate start indices *after* padding deletion (if unpad is False, these are 1)
        # Note: We still iterate over the whole sheet to find text to strip, 
        # but the strip operation only happens before row/col deletion.
        # However, for efficiency, we can focus the iteration on non-padded areas.
        
        # --- Text Stripping (before deletion) ---
        if strip_text:
            # Only iterate over the part of the sheet that is NOT padding, 
            # to avoid wasting time on empty cells in padding rows/cols.
            start_row_for_strip = padding_info['rows'] + 1
            start_col_for_strip = padding_info['cols'] + 1
            
            for row in ws.iter_rows(min_row=start_row_for_strip, min_col=start_col_for_strip):
                for cell in row:
                    # Apply stripping only to text values (data_type 's')
                    if cell.data_type == 's' and isinstance(cell.value, str):
                        original_value = cell.value
                        stripped_value = original_value.strip()
                        if stripped_value != original_value:
                            cell.value = stripped_value

        # --- Apply Unpadding (Row/Column Deletion) ---
        rows_to_delete = padding_info['rows']
        cols_to_delete = padding_info['cols']

        if rows_to_delete > 0:
            # openpyxl delete_rows(idx, amount). idx=1 means delete from the start.
            ws.delete_rows(1, rows_to_delete)
        if cols_to_delete > 0:
            ws.delete_cols(1, cols_to_delete)

    
    # 4. Third Pass (FIX): Rewrite Formulas (after deletion and using the global map)
    if unpad:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # After deletion, iteration starts from 1,1
            for row in ws.iter_rows():
                for cell in row:
                    # Only process formula cells ('f' for formula)
                    if cell.data_type == 'f' and isinstance(cell.value, str): 
                        # Crucially, we pass the GLOBAL padding map here!
                        cell.value = update_cross_sheet_formula(
                            cell.value, 
                            sheet_name, 
                            all_sheets_padding_map
                        )

    # 5. Save the modified workbook
    try:
        wb.save(outname)
        logger.info(f"Successfully processed {os.path.basename(filename)}.")
        return True
    except Exception as e:
        logger.error(f"Error saving file {os.path.basename(outname)}: {e}")
        return False

###############################################################################
# Pandas worksheet handling for unpadding and text stripping
###############################################################################

def get_padding_info_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Returns padding info of a DataFrame

    Args
        df : the dataframe to unpad

    Returns:
        the unpadded DataFrame
    

    """
    x, y = 0, 0
    while df.iloc[x].isna().all():
        x += 1
    while df.iloc[:, y].isna().all():
        y += 1
    return x, y

def unpad_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Unpads a DataFrame

    Args
        df : the dataframe to unpad

    Returns:
        the unpadded dataframe
    

    """
    x, y = get_padding_info_df(df)
    return df.iloc[x:, y:]

def strip_text_df(df: pd.DataFrame) -> pd.DataFrame:
    t = df.copy()
    stripvalue = lambda x: x if not isinstance(x, str) else x.strip()
    return t.map(stripvalue)

def unpad_strip_xls_file(filename: str, outname: str, unpad: bool, strip_text: bool) -> None:
    if not os.path.exists(filename):
        logger.error(f"File not found: {filename}")
        return False
        
    logger.info(f"Processing: {os.path.basename(filename)} (Unpad: {unpad}, Strip: {strip_text})")
    
    try:
        dfs = pd.read_excel(filename, sheet_name=None)
    except Exception as e:
        logger.error(f"Error loading file {filename}: {e}")
        sh.copy(filename, outname) # Copy source to destination for safety
        return False
    for name, df in dfs.items():
        if unpad:
            df = unpad_df(df)
        if strip_text:
            df = strip_text_df(df)
        dfs[name] = df
    with pd.ExcelWriter(outname, engine='xlwt') as writer:
        for name, df in dfs.items():
            df.to_excel(writer, sheet_name=name, index=False)

def unpad_strip_csv_file(filename: str, outname: str, unpad: bool, strip_text: bool, sep=",") -> None:
    if not os.path.exists(filename):
        logger.error(f"File not found: {filename}")
        return False
        
    logger.info(f"Processing: {os.path.basename(filename)} (Unpad: {unpad}, Strip: {strip_text})")
    
    try:
        df = pd.read_csv(filename, sep=sep)
    except Exception as e:
        logger.error(f"Error loading file {filename}: {e}")
        sh.copy(filename, outname) # Copy source to destination for safety
        return False
    if unpad:
        df = unpad_df(df)
    if strip_text:
        df = strip_text_df(df)
    df.to_csv(outname, index=False, sep=sep)

###############################################################################
# Format-agnostic unpadding and text stripping and checking
###############################################################################

def unpad_strip_file(source_path, dest_path, ext, unpad, strip_text):
    if ext == "xlsx":
        # Process Excel files
        unpad_strip_xlsx_file(source_path, dest_path, unpad, strip_text)
    else:
        if ext == "xls":
            unpad_strip_xls_file(source_path, dest_path, unpad, strip_text)
        elif ext in STRICT_SEP_EXTENSIONS:
            unpad_strip_csv_file(source_path, dest_path, unpad, strip_text, sep=EXT_TO_SEP[ext])
        else:
            # Copy other files directly
            try:
                sh.copy2(source_path, dest_path)
            except Exception as e:
                logger.error(f"Error copying non-Excel file {source_path}: {e}")
                
def unpad_strip_recursively(source_fol: str, dest_fol: str, unpad: bool, strip_text: bool):
    """Recursively processes all tabular data files in a folder."""
    source_fol = os.path.abspath(source_fol)
    dest_fol = os.path.abspath(dest_fol)
    
    if not os.path.exists(source_fol):
        logger.error(f"Source folder not found: {source_fol}")
        return

    os.makedirs(dest_fol, exist_ok=True)
    
    logger.info(f"--- Starting Folder Process: {os.path.basename(source_fol)} to {os.path.basename(dest_fol)}/ ---")

    for fol, subfols, files in os.walk(source_fol):
        # Create corresponding destination folder
        correspfol = fol.replace(source_fol, dest_fol)
        for subfol in subfols:
            os.makedirs(os.path.join(correspfol, subfol), exist_ok=True)
        
        for file in files:
            ext = os.path.splitext(file.lower())[1][1:]
            logger.info(f"ext is {ext}")
            if ext in TABULAR_EXTENSIONS:
                source_path = os.path.join(fol, file)
                dest_path = os.path.join(correspfol, file)
                logger.info(f"unpadding and/or stripping {source_path}")
                unpad_strip_file(source_path, dest_path, ext, unpad, strip_text)
                
    logger.info(f"--- Folder Process Complete: {os.path.basename(dest_fol)} ---")


def check_xls_file(filename: str, check_padding: bool, check_strip: bool) -> Optional[Dict[str, Any]]:
    """
    Checks a single .xls file for padding and/or unstripped text.
    
    Returns:
        A dictionary of issues found, or None if no issues found or file is not Excel/error occurred.
    """
    if not filename.lower().endswith(".xlsx"):
        return None
        
    try:
        dfs = pd.read_excel(filename, sheet_name=None)
    except Exception:
        logger.error(f"Could not load file {filename}. Skipping check.")
        return None
        
    issues = {'padding_found': False, 'strip_issues': False, 'details': {}}
    
    for name, df in dfs.items():
        sheet_issues = {'padding': (0, 0), 'strip_cells': []}
        
        # 1. Check Padding
        if check_padding:
            rows_to_delete, cols_to_delete = get_padding_info_df(df)
            if rows_to_delete > 0 or cols_to_delete > 0:
                issues['padding_found'] = True
                sheet_issues['padding'] = (rows_to_delete, cols_to_delete)
        
        # 2. Check Text Stripping
        if check_strip:
            # We only need to check the remaining cells (i.e., skipping any padded area)
            start_row = sheet_issues['padding'][0] + 1
            start_col = sheet_issues['padding'][1] + 1
            
            for n, row in df.iloc[start_row:, start_col:].iter_rows():
                for m, cell in enumerate(row):
                    # Check for text (data_type 's') that starts or ends with space
                    if isinstance(cell, str):
                        if cell.value.startswith(' ') or cell.value.endswith(' '):
                            issues['strip_issues'] = True
                            sheet_issues['strip_cells'].append([n + start_row, m + start_col])
            
        if issues['padding_found'] or issues['strip_issues']:
            issues['details'][name] = sheet_issues
            
    return issues if issues['padding_found'] or issues['strip_issues'] else None

def check_xlsx_file(filename: str, check_padding: bool, check_strip: bool) -> Optional[Dict[str, Any]]:
    """
    Checks a single .xlsx file for padding and/or unstripped text.
    
    Returns:
        A dictionary of issues found, or None if no issues found or file is not Excel/error occurred.
    """
    if not filename.lower().endswith(".xlsx"):
        return None
        
    try:
        wb = load_workbook(filename)
    except Exception:
        logger.error(f"Could not load file {filename}. Skipping check.")
        return None
        
    issues = {'padding_found': False, 'strip_issues': False, 'details': {}}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_issues = {'padding': (0, 0), 'strip_cells': []}
        
        # 1. Check Padding 
        if check_padding:
            rows_to_delete, cols_to_delete = get_padding_info_ws(ws)
            if rows_to_delete > 0 or cols_to_delete > 0:
                issues['padding_found'] = True
                sheet_issues['padding'] = (rows_to_delete, cols_to_delete)
        
        # 2. Check Text Stripping 
        if check_strip:
            # We only need to check the remaining cells (i.e., skipping any padded area)
            start_row = sheet_issues['padding'][0] + 1
            start_col = sheet_issues['padding'][1] + 1
            
            for row in ws.iter_rows(min_row=start_row, min_col=start_col):
                for cell in row:
                    # Check for text (data_type 's') that starts or ends with space
                    if cell.data_type == 's' and isinstance(cell.value, str):
                        if cell.value.startswith(' ') or cell.value.endswith(' '):
                            issues['strip_issues'] = True
                            sheet_issues['strip_cells'].append(cell.coordinate)
            
        if issues['padding_found'] or issues['strip_issues']:
            issues['details'][sheet_name] = sheet_issues
            
    return issues if issues['padding_found'] or issues['strip_issues'] else None

def check_csv_file(filename: str, check_padding: bool, check_strip: bool, sep=",") -> Optional[Dict[str, Any]]:
    """
    Checks a single csv file for padding and/or unstripped text.
    
    Returns:
        A dictionary of issues found, or None if no issues found or file is not Excel/error occurred.
    """
    ext = os.path.splitext(filename.lower())[1][1:]
    if ext not in STRICT_SEP_EXTENSIONS:
        return None
        
    try:
        df = pd.read_csv(filename, sep=sep)
    except Exception:
        logger.error(f"Could not load file {filename}. Skipping check.")
        return None
        
    issues = {'padding_found': False, 'strip_issues': False, 'details': {"only_sheet": {'padding': (0, 0), 'strip_cells': []}}}
    
    # 1. Check Padding
    if check_padding:
        rows_to_delete, cols_to_delete = get_padding_info_df(df)
        if rows_to_delete > 0 or cols_to_delete > 0:
            issues['padding_found'] = True
            issues["details"]["only_sheet"]['padding'] = (rows_to_delete, cols_to_delete)
    
    # 2. Check Text Stripping
    if check_strip:
        # We only need to check the remaining cells (i.e., skipping any padded area)
        start_row = issues["details"]["only_sheet"]['padding'][0] + 1
        start_col = issues["details"]["only_sheet"]['padding'][1] + 1
        
        for n, row in df.iloc[start_row:, start_col:].iterrows():
            for m, cell in enumerate(row):
                # Check for text (data_type 's') that starts or ends with space
                if isinstance(cell, str):
                    if cell.value.startswith(' ') or cell.value.endswith(' '):
                        issues['strip_issues'] = True
                        issues["details"]["only_sheet"]['strip_cells'].append([n + start_row, m + start_col])
        
    return issues if issues['padding_found'] or issues['strip_issues'] else None

def check_file(full_path, ext, check_padding, check_strip, folder_path=None):
    if ext in STRICT_SEP_EXTENSIONS:
        issues = check_csv_file(full_path, check_padding, check_strip, sep=EXT_TO_SEP[ext])
    elif ext == "xlsx":
        logger.info("running check_xlsx")
        issues = check_xlsx_file(full_path, check_padding, check_strip)
    elif ext == "xls":
        issues = check_xls_file(full_path, check_padding, check_strip)
    else:
        issues = False
    if issues:
        relative_path = os.path.relpath(full_path, folder_path) if folder_path else full_path
        logger.warning(f"\n[ISSUE FOUND]: {relative_path}")
        
        if issues['padding_found']:
            logger.warning("  * Padding Found (Run 'unpad' or 'clean' to fix):")
            for sheet, detail in issues['details'].items():
                if detail['padding'][0] > 0 or detail['padding'][1] > 0:
                    logger.warning(f"    - Sheet '{sheet}': {detail['padding'][0]} row(s), {detail['padding'][1]} col(s)")
                    
        if issues['strip_issues']:
            logger.warning("  * Unstripped Text Found (Run 'strip-text' or 'clean' to fix):")
            for sheet, detail in issues['details'].items():
                if detail['strip_cells']:
                    # Only show the first few cells to keep the output clean
                    coords = detail['strip_cells'][:5]
                    more = f"... (+{len(detail['strip_cells']) - 5} more)" if len(detail['strip_cells']) > 5 else ""
                    logger.warning(f"    - Sheet '{sheet}': e.g., {', '.join(coords)}{more}")
        logger.debug("done")
    return issues

def check_recursively(folder_path: str, check_padding: bool, check_strip: bool):
    """
    Recursively checks all Excel files in a folder for issues and prints a report.
    """
    if not os.path.isdir(folder_path):
        logger.error(f"Folder not found: {folder_path}")
        return

    logger.info(f"--- Starting Recursive Check in: {os.path.basename(folder_path)} ---")
    
    found_issues = False
    
    for fol, _, files in os.walk(folder_path):
        for file in files:
            ext = os.path.splitext(file.lower())[1][1:]
            if ext in TABULAR_EXTENSIONS:
                full_path = os.path.join(fol, file)
                issues_in_file = check_file(full_path, ext, check_padding, check_strip, folder_path=folder_path)
                if issues_in_file:
                    found_issues = True

    if not found_issues:
        logger.info("\n--- Check Complete: No issues found in any tabular data file. ---")
    else:
        logger.info("\n--- Check Complete: Issues reported above. ---")
        
 ###############################################################################
 # Table manipulations
 ###############################################################################
       
def read_sheets(file, frmt=None):
    """
    Reads an Excel (xlsx/xls) or CSV file into a pandas DataFrame.
    """
    file_lower = file.lower()
    if frmt is None:
        frmt = os.path.splitext(file_lower)[1][1:]
        if frmt not in TABULAR_EXTENSIONS:
            raise InvalidFileFormatError("Only csv/tsv, xlsx, and xls can be processed")
    
    # Determine the read function
    if frmt in STRICT_SEP_EXTENSIONS:
        sep = EXT_TO_SEP[frmt]
        # Read csv
        fname = os.path.splitext(os.path.split(file)[1])[0]
        sheets = {fname: {"df": pd.read_csv(file, sep=sep)}}
        # Re-read the file to capture the actual first row as the list of column names
        try:
            # Set header=None and read only the first row
            columns_df = pd.read_csv(file, nrows=1, header=None, sep=sep)
            # Extract the original column names (potentially non-unique)
            sheets[fname]["columns"] = columns_df.iloc[0].fillna(value="").astype(str).tolist()
        except Exception as e:
            logger.warning(f"Could not read first row for original column names in {file}: {e}")
            sheets[fname]["columns"] = sheets[fname]["df"].columns.astype(str).tolist() # Fallback to pandas detected headers
    elif frmt in PROCESS_EXTENSIONS:
        sheets = {k: {"df": v} for k, v in pd.read_excel(file, sheet_name=None).items()}
        for sheet_name, dict_ in sheets.items():
            try:
                # Set header=None and read only the first row
                columns_df = pd.read_excel(file, nrows=1, header=None, sheet_name=sheet_name)
                # Extract the original column names (potentially non-unique)
                sheets[sheet_name]["columns"] = columns_df.iloc[0].fillna(value="").astype(str).tolist()
            except Exception as e:
                logger.warning(f"Could not read first row for original column names in {file}: {e}")
                sheets[sheet_name]["columns"] = sheets[sheet_name]["df"].columns.astype(str).tolist() # Fallback to pandas detected headers
    else:
         raise InvalidFileFormatError(f"Unsupported format for reading: {frmt}")

    return sheets, frmt

def check_multitable_df(df, file, sheet=None):
    """
    Checks if a DataFrame contains multiple tables separated by fully empty columns (NaNs).
    Returns True if multiple tables are found, False otherwise.
    """
    empty_cols = []
    
    for n, col in enumerate(df.columns):
        # Check if the entire column consists of NaN values
        if df[col].isna().all():
            empty_cols.append(n)

    empty_rows = []
    for n, i in enumerate(df.index):
        if df.iloc[n].isna().all():
            empty_rows.append(n)
    
    if any([empty_cols, empty_rows]):
        columns_bit = f"columns at indices {empty_cols} " if empty_cols else ""
        rows_bit = f"rows at indices {empty_rows} " if empty_rows else ""
        bit = f"{columns_bit}{'and ' if empty_cols and empty_rows else ''}{rows_bit}"
        sheet_bit = f", sheet {sheet}" if sheet else ""
        logger.warning(f"Multiple tables issue in {file}{sheet_bit}: {bit}are empty and suggest a table split.")
        return True
    return False

def check_multitable_file(fname, ext):
    if ext in STRICT_SEP_EXTENSIONS:
        check_multitable_df(pd.read_csv(fname, sep=EXT_TO_SEP[ext]), fname)
    if ext in PROCESS_EXTENSIONS:
        dfs = pd.read_excel(fname, sheet_name=None)
        for sheet_name, df in dfs.items():
            check_multitable_df(df, fname, sheet=sheet_name)
            
def check_multitable_recursively(folder):
    for folder, subfolders, files in os.walk(folder):
        for file in files:
            ext = os.path.splitext(file.lower())[1][1:]
            if ext in TABULAR_EXTENSIONS:
                fname = os.path.join(folder, file)
                check_multitable_file(fname, ext)

def _safe_sheet_name(name: Any) -> str:
    """Sanitizes a string to be a valid, truncated Excel sheet name."""
    name_str = str(name)
    # Invalid characters: \ / ? * [ ] : 
    name_str = name_str.replace('[', '').replace(']', '').replace(':', '').replace('/', '').replace('\\', '').replace('?', '').replace('*', '').replace(' ', '_')
    # Limit length to 31 characters
    return name_str[:31]

def _get_excel_writer_engine(out_format: str) -> str:
    """Returns the correct Pandas ExcelWriter engine based on output format."""
    if out_format == 'xlsx':
        return 'xlsxwriter'
    elif out_format == 'xls':
        # Requires the 'xlwt' library
        return 'xlwt'
    else:
        raise InvalidFileFormatError(f"Unsupported Excel format for writing: {out_format}")

def check_and_clean_folderpath(path):
    assert not os.path.splitext(path)[1], f"It looks like you provided a filepath ({path}), while the code was expecting a folder path."
    if path.endswith(os.path.sep):
        path = f"{path}{os.path.sep}"
    return path

def write_tables(tables_per_sheet, source_file, out_format, destination, inplace, operation, operation_name):
    basename = os.path.splitext(source_file)[0]
    folder_name = os.path.split(source_file)[0]
    if destination:
        os.makedirs(destination, exist_ok=True)
    
    if out_format in PROCESS_EXTENSIONS:
        if inplace:
            out_file = source_file
        else:
            if destination:
                out_file = os.path.join(destination, os.path.split(f"{basename}_{operation}.{out_format}")[1])
            else:
                out_file = f"{basename}_{operation}.{out_format}"
        
        # Ensure the source_filename is not too long
        if len(out_file) > 218: 
             logger.warning(f"Output source_filename {out_file} is too long, truncating.")
             out_file = os.path.join(destination, f"{basename[:210]}_{operation}.{out_format}")

        # 4. Write Output (XLSX/XLS)
        try:
            engine = _get_excel_writer_engine(out_format)
            with pd.ExcelWriter(out_file, engine=engine) as writer:
                for sheet_name, tables in tables_per_sheet.items():
                    for k, table in tables.items():
                        new_sheet_name = _safe_sheet_name(f"{sheet_name}_{k}")
                        table.to_excel(writer, sheet_name=new_sheet_name, index=False, header=True)
            logger.info(f"Successfully {operation_name} from {source_file} to {out_format} sheets in {out_file}")
        except Exception as e:
            logger.error(f"Error writing {out_format} output for {source_file}: {e}")

    elif out_format in STRICT_SEP_EXTENSIONS:
        # Write each table to a separate CSV source_file
        for sheet_name, tables in tables_per_sheet.items():
            for k, table in tables.items():
                safe_k = _safe_sheet_name(k)
                if destination:
                    out_file = os.path.join(destination, f"{sheet_name}_{operation}_{safe_k}.{out_format}")
                else:
                    out_file = os.path.join(folder_name, f"{basename}_{operation}_{safe_k}.{out_format}")
                table.to_csv(out_file, index=False, header=True, sep=EXT_TO_SEP[out_format])
        logger.info(f"Successfully {operation_name} from {source_file} into multiple {out_format.upper()}")
        if inplace:
            os.remove(source_file)
    else:
        logger.error(f"Unsupported output format: {out_format}")

def vsplit_tables(file, in_format=None, out_format=None, inplace=False, destination=None):
    """
    Splits a file containing multiple tables separated by empty columns into 
    individual sheets (if XLSX/XLS) or separate CSV files.
    """
    try:
        sheets, in_format = read_sheets(file, frmt=in_format)
    except InvalidFileFormatError as e:
        logger.error(f"Skipping split for {file}: {e}")
        return
    if destination:
        destination = check_and_clean_folderpath(destination)
        
    tables_per_sheet: Dict[str, dict]  = {}
    multi_tables = False
    for sheet_name, data in sheets.items():
        df, original_columns = data["df"], data["columns"]
        start = 0
        tables: Dict[str, pd.DataFrame] = {}
        
        # 1. Identify table boundaries (columns that are all NaN)
        empty_cols_indices = [n for n, col in enumerate(df.columns) if df[col].isna().all()]
        if not empty_cols_indices:
            logger.info(f"No multiple tables found in sheet {sheet_name} in {file} to split.")
            tables_per_sheet[sheet_name] = {sheet_name: df}
            continue
        multi_tables = True
        table_boundaries = empty_cols_indices + [len(df.columns)]
        
        # 2. Extract tables
        for n, block_end in enumerate(table_boundaries):
            end = block_end
            if end > start:  # Extract table if block is not empty
                table = df.iloc[:, start:end].copy()
                # logger.info(table.iloc[:3])
                if not any(original_columns[start:end]):  # empty header line
                    table = table.iloc[1:]
                    key = n + 1
                elif original_columns[start] and not any(original_columns[start+1:end]):  # table title and column headers
                    key = df.columns[start]
                    table.columns = table.iloc[0].tolist()
                    table = table.iloc[1:]
                else:
                    table.columns = original_columns[start:end]
                    key = n + 1
                tables[key] = table
            start = block_end + 1
        tables_per_sheet[sheet_name] = tables
    if not multi_tables:
        logger.info(f"No multiple tables at all in {file}")
        if destination:  # no point in re-writing the file
            sh.copy2(file, destination)
        return
    out_format = in_format if out_format is None else out_format
    write_tables(tables_per_sheet, file, out_format, destination, inplace, "vsplit", "split tables vertically")


def vsplit_into_two_colum_tables(file, in_format=None, out_format=None, inplace=False, destination=None):
    """
    Splits a file containing multiple tables into two-column (X, Y) pairs.
    """
    try:
        sheets, in_format = read_sheets(file, frmt=in_format)
    except InvalidFileFormatError as e:
        logger.error(f"Skipping 2-column split for {file}: {e}")
        return
    
    tables_per_sheet: Dict[str, dict]  = {}
    for sheet_name, data in sheets.items():
        df, original_columns = data["df"], data["columns"]
        tables: Dict[str, pd.DataFrame] = {}
        
        # 1. Identify table boundaries (columns that are all NaN)
        empty_cols_indices = [n for n, col in enumerate(df.columns) if df[col].isna().all()]
        
        # Define block start and end indices
        block_starts = [0] + [i + 1 for i in empty_cols_indices]
        block_ends = empty_cols_indices + [len(df.columns)]
        
        # Filter for valid blocks (start < end)
        blocks: List[Tuple[int, int]] = [(s, e) for s, e in zip(block_starts, block_ends) if s < e]
    
        # 2. Extract tables in (X, Y) pairs
        for start, end in blocks:
            block_df = df.iloc[:, start:end]
            
            # The first column of the block is the X-axis (start).
            if block_df.shape[1] > 1:
                x_col_index = start
                x_col_name_original = original_columns[x_col_index]
                
                # Create (X, Y) pairs for all Y columns in this block
                for y_index_in_block in range(1, block_df.shape[1]):
                    y_col_index = start + y_index_in_block
                    
                    y_col_name_original = original_columns[y_col_index] 
                    y_col_name_unique = df.columns[y_col_index] # Unique Pandas name for key
                    
                    # Create the two-column table
                    table = df.iloc[:, [x_col_index, y_col_index]].copy()
                    
                    # Use original column names for the headers
                    table.columns = [x_col_name_original, y_col_name_original]
                    
                    # Use the unique Pandas Y column name as the key
                    key = y_col_name_unique
                    tables[key] = table
        tables_per_sheet[sheet_name] = tables

    # 3. Handle Output Path
    out_format = in_format if out_format is None else out_format
    write_tables(tables_per_sheet, file, out_format, destination, inplace, "2col_split", "split into 2 column tables")
    
def hsplit_tables(file, in_format=None, out_format=None, inplace=False, destination=None):
    """
    Splits a file containing multiple tables separated by empty columns into 
    individual sheets (if XLSX/XLS) or separate CSV files.
    """
    try:
        sheets, in_format = read_sheets(file, frmt=in_format)
    except InvalidFileFormatError as e:
        logger.error(f"Skipping split for {file}: {e}")
        return
    if destination:
        destination = check_and_clean_folderpath(destination)
        
    tables_per_sheet: Dict[str, dict]  = {}
    multi_tables = False
    for sheet_name, data in sheets.items():
        df, original_columns = data["df"], data["columns"]
        if sum(bool(x) for x in original_columns) == 1:  # header was actually table title
            df.columns = df.iloc[0].fillna(value="").astype(str).tolist()
            df.iloc[0] = original_columns
        start = 0
        tables: Dict[str, pd.DataFrame] = {}
        
        # 1. Identify table boundaries (columns that are all NaN)
        empty_rows_indices = [n for n, col in enumerate(df.index) if df.iloc[n].isna().all()]
        if not empty_rows_indices:
            logger.info(f"No multiple tables found in sheet {sheet_name} in {file} to split.")
            tables_per_sheet[sheet_name] = {sheet_name: df}
            continue
        multi_tables = True
        table_boundaries = empty_rows_indices + [len(df.index)]
        
        # 2. Extract tables
        for n, block_end in enumerate(table_boundaries):
            key = False
            end = block_end
            if end > start:  # Extract table if block is not empty
                table = df.iloc[start:end, :].copy()
                table.columns = original_columns
                row0, row1 = table.iloc[0], table.iloc[1]
                if sum(bool(pd.notna(x)) for x in row0) == 1:  # only 1 notna => table title
                    idx = [n for n, x in enumerate(row0) if pd.notna(x)][0]
                    key = row0.iloc[idx]
                    table = table.iloc[1:]
                    row0 = row1
                key = n + 1 if not key else key
                if all(isinstance(x, str) or pd.isna(x) for x in row0):  # not data
                    table.columns = [str(x) if pd.notna(x) else "" for x in table.iloc[0]]
                    table = table.iloc[1:]
                tables[key] = table
            start = block_end + 1
        if not tables:
            logger.info(f"No multiple tables found in sheet {sheet_name} in {file} to split.")
            tables_per_sheet[sheet_name] = {sheet_name: df}
        else:
            multi_tables = True
        tables_per_sheet[sheet_name] = tables
    if not multi_tables:
        logger.info(f"No multiple tables at all in {file}")
        if destination:  # no point in re-writing the file
            sh.copy2(file, destination)
        return
    
    out_format = in_format if out_format is None else out_format
    write_tables(tables_per_sheet, file, out_format, destination, inplace, "hsplit", "split tables horizontally")

def convert_file(file, out_format=None, destination=None, inplace=False, sep=None):
    if out_format is None:
        raise ValueError("You must select an output format")
    if destination:
        os.makedirs(destination, exist_ok=True)
    ext = os.path.splitext(file.lower())[1][1:]
    folder_path, fname = os.path.split(file)
    basename = fname[:-(len(ext)+1)]
    if destination is None:
        destination = folder_path
    if out_format == ext:
        if inplace:
            return
        else:
            sh.copy2(file, os.path.join(destination, fname))
    if ext in WIDE_SEP_EXTENSIONS:
        if sep is None:
            EXT_TO_SEP[ext]
        dfs = {basename: pd.read_csv(file, sep=sep, header=None)}
    elif ext in PROCESS_EXTENSIONS:
        dfs = pd.read_excel(file, sheet_name=None, header=None)
    out_folder = folder_path if inplace else destination
    if out_format in STRICT_SEP_EXTENSIONS:
        for sheet_name, df in dfs.items():
            addendum = "" if len(dfs) == 1 else f"_{sheet_name}"
            df.to_csv(os.path.join(out_folder, f"{basename}{addendum}.{out_format}"), header=False, index=False, sep=EXT_TO_SEP[out_format])
    elif out_format in PROCESS_EXTENSIONS:
        for sheet_name, df in dfs.items():
            engine = _get_excel_writer_engine(out_format)
            addendum = "" if len(dfs) == 1 else f"_{sheet_name}"
            t = os.path.join(out_folder, f"{basename}{addendum}.{out_format}")
            logger.info(t)
            with pd.ExcelWriter(os.path.join(out_folder, f"{basename}{addendum}.{out_format}"), engine=engine) as writer:
                for sheet_name, df in dfs.items():
                    new_sheet_name = _safe_sheet_name(sheet_name)
                    df.to_excel(writer, sheet_name=new_sheet_name, index=False, header=False)
    else:
        raise InvalidFileFormatError("Unsupported output format {out_format}")
    if inplace:
        os.remove(file)

def detect_table(bool_df, point):
    x, y = point
    edges = np.array([[x, x], [y, y]])
    nrows, ncols = bool_df.shape
    maxs = ncols -1, nrows -1 
    while True:
        changed = False
        for n in range(2):  # x or y
            for m in range(2):  # first or second edge
                incr = 1 if m else -1
                new_edges = edges.copy()
                if 0 <= new_edges[n,m] + incr <=  maxs[n]:
                    new_edges[n, m] += incr
                else:
                    continue
                if n:
                    non_empty = bool_df.iloc[new_edges[1,m], new_edges[0,0]:new_edges[0,1] + 1].any()
                else:
                    non_empty = bool_df.iloc[new_edges[1,0]:new_edges[1,1]+1, new_edges[0,m]].any()
                if non_empty:
                    changed = True
                    edges = new_edges
        if not changed:
            break
    return edges
            
def slice_table(df, edges):
    return df.iloc[edges[1,0]: edges[1,1] + 1, edges[0,0]: edges[0,1] + 1].copy()

def point_in_table(edges, point, padding=False, nrows=None, ncols=None):
    edges_cp = edges.copy()
    if padding:
        if nrows is None or ncols is None:
            raise ValueError("if you want padding, you must provide nrows and ncols")
        if edges_cp[0, 0]:
            edges_cp[0,0] -= 1
        if edges_cp[1, 0]:
            edges_cp[1, 0] -= 1
        if edges_cp[0,1] < ncols - 1:
            edges_cp[0,1] += 1
        if edges_cp[1,1] < nrows - 1:
            edges_cp[1,1] += 1
    return edges_cp[0,0] <= point[0] <= edges_cp[0,1] and edges_cp[1,0] <= point[1] <= edges_cp[1,1]

def point_in_any_table(edges_list, point, padding=False, nrows=None, ncols=None):
    for edges in edges_list:
        in_table = point_in_table(edges, point, padding=padding, nrows=nrows, ncols=ncols)
        if in_table:
            return True
    return False

def detect_table_edges(bool_df):
    nrows, ncols = bool_df.shape
    coords = product(range(ncols), range(nrows))
    seen = set()
    table_edges = []
    for point in coords:
        if bool_df.iloc[point[1], point[0]] and (point not in seen) and not point_in_any_table(table_edges, point, padding=True, nrows=nrows, ncols=ncols):
            new_table_edges = detect_table(bool_df, point)
            table_edges.append(new_table_edges)
        seen.add(point)
    return table_edges

def get_tables_df(df):
    bool_df = df.notna()
    table_edges = detect_table_edges(bool_df)
    tables = {}
    if len(table_edges) == 1:
        return {"": df}
    for n, edges in enumerate(table_edges):
        key = False
        table = slice_table(df, edges)
        row0, row1 = table.iloc[0], table.iloc[1]
        if sum(bool(pd.notna(x)) for x in row0) == 1:  # only 1 notna => table title
            idx = [n for n, x in enumerate(row0) if pd.notna(x)][0]
            key = row0.iloc[idx]
            table = table.iloc[1:]
            row0 = row1
        key = n + 1 if not key else key
        if all(isinstance(x, str) or pd.isna(x) for x in row0):  # not data
            table.columns = [str(x) if pd.notna(x) else "" for x in table.iloc[0]]
            table = table.iloc[1:]
        tables[key] = table
    return tables

def split_tables_file(file, in_format=None, out_format=None, inplace=False, destination=None):
    try:
        sheets, in_format = read_sheets(file, frmt=in_format)
    except InvalidFileFormatError as e:
        logger.error(f"Skipping split for {file}: {e}")
        return
    if destination:
        destination = check_and_clean_folderpath(destination)
    
    tables_per_sheet = {}
    for sheet_name, data in sheets.items():
        df = data["df"]
        tables_per_sheet[sheet_name] = get_tables_df(df)
    if all([len(v) == 1 for k, v in tables_per_sheet.items()]):
        if destination:
            fname =os.path.split(file)[1]
            sh.copy2(file, os.path.join(destination, fname))
    out_format = in_format if out_format is None else out_format
    write_tables(tables_per_sheet, file, out_format, destination, inplace, "splitall", "split all tables")
    
def process_recursively(path: str, file_func: Callable[..., None], destination=None,
                        out_format=None, inplace=False, format_to_process=None,
                        **kwargs) -> None:
    """
    Recursively processes all supported tabular files (.xlsx, .xls, .csv) 
    in a directory or processes a single file, applying the provided file_func.

    Args:
        path: Path to the single file or root directory.
        file_func: The function to apply to each file path.
        **kwargs: Additional keyword arguments passed to file_func.
    """
    path = os.path.abspath(path)
    is_dir: bool = os.path.isdir(path)
    is_file: bool = os.path.isfile(path)

    if is_dir:
        source_fol: str = path
        for fol, subfols, files in os.walk(source_fol):
            if destination:
                correspfol = fol.replace(source_fol, destination)
                for subfol in subfols:
                    os.makedirs(os.path.join(correspfol, subfol), exist_ok=True)
            for file in files:
                file_path: str = os.path.join(fol, file)
                ext = os.path.splitext(file)[1][1:]
                to_process = ext == format_to_process if format_to_process else file.lower().endswith(TABULAR_EXTENSIONS)
                if to_process:
                    logger.info(f"Processing file: {file_path}")
                    try:
                        dest_path = None if destination is None else correspfol
                        file_func(file_path, destination=dest_path,
                                  out_format=out_format, inplace=inplace,
                                  **kwargs)
                    except Exception as e:
                        logger.error(f"Failed to process {file_path}: {e}")
                elif not inplace:
                    sh.copy2(file_path, os.path.join(correspfol, file))
    elif is_file:
        if not path.lower().endswith(TABULAR_EXTENSIONS):
             logger.error(f"File {path} is not a supported tabular format ({', '.join(TABULAR_EXTENSIONS)}).")
             return
        logger.info(f"Processing single file: {path}")
        try:
            file_func(path, destination=destination,
                      out_format=out_format, inplace=inplace,
                      **kwargs)
        except Exception as e:
            logger.error(f"Failed to process {path}: {e}")
    else:
        logger.error(f"Invalid path: {path} is neither a file nor a directory.")
     
    
# =============================================================================
# CLI Implementation
# =============================================================================

def check_command(args):
    """
    Function to handle the 'process' command logic.
    Placeholder for your actual processing code.
    """
    if not os.path.exists(args.source):
        raise FileNotFoundError(f"Your source {args.source} does not exist!!")
    if os.path.isfile(args.source):
        ext = os.path.splitext(args.source.lower())[1][1:]
    # Logic for the mutually exclusive options
    if any([args.strip_only, args.unpad_only, args.strip_unpad]):
        check_padding = False if args.strip_only else True
        check_strip = False if args.unpad_only else True
        if os.path.isfile(args.source):
            check_file(args.source, ext, check_padding, check_strip)
        elif os.path.isdir(args.source):
            check_recursively(args.source, check_padding, check_strip)
    elif args.multi_table:
        if os.path.isfile(args.source):
            check_multitable_file(args.source, ext)
        elif os.path.isdir(args.source):
            check_multitable_recursively(args.source)

def process_command(args):
    """
    Function to handle the 'check' command logic.
    Placeholder for your actual checking code.
    """
    if not os.path.exists(args.source):
        FileNotFoundError(f"Your source {args.source} does not exist!!")
    if os.path.isfile(args.source):
        ext = os.path.splitext(args.source.lower())[1][1:]
    # Logic for the mutually exclusive options
    if any([args.strip_only, args.unpad_only, args.strip_unpad]):
        if args.out_format:
            logger.info(f"Ignoring '--out-format {args.out_format}' because the processing selected always maintains format")
        dest = args.source if args.inplace else args.destination
        unpad = False if args.strip_only else True
        strip_text = False if args.unpad_only else True
        if os.path.isfile(args.source):
            unpad_strip_file(args.source, dest, ext, unpad, strip_text)
        elif os.path.isdir(args.source):
            unpad_strip_recursively(args.source, dest, unpad, strip_text)
    else:
        if args.vsplit_tables:
            split_func = vsplit_tables
        elif args.vsplit_into_two_columns_tables:
            split_func = vsplit_into_two_colum_tables
        elif args.hsplit_tables:
            split_func = hsplit_tables
        elif args.split_all_tables:
            split_func = split_tables_file
    
        if os.path.isfile(args.source):
            split_func(args.source, in_format=ext, out_format=args.out_format, inplace=args.inplace, destination=args.destination)
        elif os.path.isdir(args.source):
            process_recursively(args.source, split_func, out_format=args.out_format, destination=args.destination, inplace=args.inplace)

def convert_command(args):
    if not os.path.exists(args.source):
        FileNotFoundError(f"Your source {args.source} does not exist!!")
    if os.path.isfile(args.source):
        convert_file(args.source, out_format=args.out_format, destination=args.destination, inplace=args.inplace, sep=args.sep)
    elif os.path.isdir(args.source):
        process_recursively(args.source, convert_file, destination=args.destination, inplace=args.inplace,
                            out_format=args.out_format, format_to_process=args.in_format, sep=args.sep)

def cli():
    """Configures and runs the command line interface."""
    parser = argparse.ArgumentParser(
        description=(
            "A utility for checking or cleaning tabular data files (unpadding, text stripping, multiple tables) "
        ),
        formatter_class=argparse.RawTextHelpFormatter
    )
    # Use add_subparsers to handle 'process' and 'check' commands
    subparsers = parser.add_subparsers(
        title='commands',
        description='valid commands',
        help='available actions',
        required=True
    )
    
        
    parser_process = subparsers.add_parser(
        'process', 
        help='Process and transform tabular data.'
    )
    parser_process.set_defaults(func=process_command)
    
    # 1. Compulsory 'source' argument
    parser_process.add_argument(
        'source',
        help='The source file or directory to process.'
    )
    
    parser_process.add_argument(
        '--out-format',
        type=str,
        dest='out_format',
        help='The output format for split operations (used only with splitting table options). You can use csv, tsv, xlsx, xls'
    )

    # 2. Mutually Exclusive Group for output location (Required for PROCESS)
    location_group_process = parser_process.add_mutually_exclusive_group(required=True)
    
    location_group_process.add_argument(
        '--inplace',
        action='store_true',
        dest='inplace',
        help='Modify the source file(s) in-place.'
    )
    location_group_process.add_argument(
        '--destination',
        type=str,
        dest='destination',
        help='The destination file or directory for the output.'
    )
    
    parser_check = subparsers.add_parser(
        'check', 
        help='Check tabular data for formatting and structure issues.'
    )
    parser_check.set_defaults(func=check_command)
    
    # 1. Compulsory 'source' argument
    parser_check.add_argument(
        'source',
        help='The source file or directory to check.'
    )
    
    
    # Mutually Exclusive Group for 'process' options
    process_group = parser_process.add_mutually_exclusive_group(required=True)
    process_group.add_argument('--strip-only', action='store_true', help='Only strip whitespace from cell contents.')
    process_group.add_argument('--unpad-only', action='store_true', help='Only unpad data to remove column spacing.')
    process_group.add_argument('--strip-unpad', action='store_true', help='Strip whitespace AND unpad data.')
    process_group.add_argument('--vsplit-tables', action='store_true', help='Split vertical multitables.')
    process_group.add_argument('--vsplit-into-two-columns-tables', action='store_true', help='Vertically split into two columns tables.')
    process_group.add_argument('--hsplit-tables', action='store_true', help='Split horizontal multitables')
    process_group.add_argument('--split-all-tables', action='store_true', help='Split all multitables')
    
    # Mutually Exclusive Group for 'check' options
    check_group = parser_check.add_mutually_exclusive_group(required=True)
    check_group.add_argument('--strip-only', action='store_true', help='Check only for cells needing strip.')
    check_group.add_argument('--unpad-only', action='store_true', help='Check only for padding issues.')
    check_group.add_argument('--strip-unpad', action='store_true', help='Check for both strip and unpad issues.')
    check_group.add_argument('--multi-table', action='store_true', help='Check for multiple tables in a single file.')

    parser_convert = subparsers.add_parser(
        'convert', 
        help='Just convert tabular data files between different extesions.'
    )
    parser_convert.set_defaults(func=convert_command)
    
    # 1. Compulsory 'source' argument
    parser_convert.add_argument(
        'source',
        help='The source file or directory to process.'
    )
    
    parser_convert.add_argument(
        '--out-format',
        type=str,
        dest='out_format',
        help='The output format for the conversion You can use csv, tsv, xlsx, xls'
    )
    
    parser_convert.add_argument(
        '--in-format',
        type=str,
        dest='in_format',
        help='The extension of files you want to process if working in a folder'
    )
    
    parser_convert.add_argument(
        'separator',
        dest="sep",
        help='The separator used (e.g. "," or ";"). Space is the default for .txt and .dat. To explicitly specify space use "\s+"'
    )

    # 2. Mutually Exclusive Group for output location (Required for PROCESS)
    location_group_convert = parser_convert.add_mutually_exclusive_group(required=True)
    
    location_group_convert.add_argument(
        '--inplace',
        action='store_true',
        dest='inplace',
        help='Modify the source file(s) in-place.'
    )
    location_group_convert.add_argument(
        '--destination',
        type=str,
        dest='destination',
        help='The destination file or directory for the output.'
    )
    # Parse arguments and call the corresponding function
    if len(sys.argv) == 1:
        parser.print_help(sys.stderr)
        sys.exit(1)

    args = parser.parse_args()
    args.func(args)



if __name__ == '__main__':
    cli()